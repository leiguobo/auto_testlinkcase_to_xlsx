# coding:utf-8
"""
@author:leiguobo
@time: 2021/06/07
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import re,os
from TestCase2Testlink.common.common1 import TestLinkOperate
# from test_to_xlsx.testcase_module import module_list

"""
使用方法：查看readme.md
"""
server = "http://xxxx/testlink/lib/api/xmlrpc/v1/xmlrpc.php"
api_key = "xxxxxx"  # testlink上获取
file_name = "testCase_Example.xlsx"  # 用例文件名
project_id = "xx"
father_id = "xxxx"

TO = TestLinkOperate(server, api_key)
'''
# 上传用例至指定用例集
TO.upload(project_id, father_id, file_name, sheet_num=0)  # sheet_num 指定上传第几个sheet页的数据
'''
# 下载指定用例集数据
TO.download(father_id)
project_name=TO.get_suites(father_id)['name']
from test_to_xlsx.gain_case_path import case_catalogue_path_last
from testlink2metersphere.testlink2metersphere2 import out_testsuite
module_list=out_testsuite

testcase_list = []  # 用例名称
# module_list = []  # 所属模块
type_list = ['functional']  # 用例类型
maintainer_list = ['test']  # 维护人
testcase_level_list = ['P0']  # 用例等级
test_way_list = ['manual']  # 测试方式
preconditions_list = []  # 前置条件
comment_list = ['none']  # 备注
actions_list = []  # 操作步骤
expectedresults_list = []  # 期望结果
data = pd.read_excel(case_catalogue_path_last,sheet_name='Sheet1',usecols=[0,1,2,3,8])
data=data.values
pattern = re.compile(r'(complex)(.*)(complex)')

excel_name=project_name+'_metersphere'+'.xlsx'
sheet_name=project_name

#筛选数据，并存储
for i in range(len(data)):
    try:
        # 用空替换\n\n
        clean_line = str(data[i]).replace('\\n\\n', '')
        # 用none替换nan
        replace_character1 = clean_line.replace('nan', "'none'").replace("&nbsp;","").replace("code","")\
            .replace("{","").replace("}","").replace("\n", "").replace("&ldquo;", "“").replace("quot",'')\
            .replace("defaultTAB","").replace("&rdquo;", "”").replace("<br />",'').replace("nbsp;",'')\
            .replace("data",'').replace('；','\n').replace(" &gt;","").replace('&;','"').replace("</p>", " ")\
            .replace("</li>","").replace("\\t","").replace("<p>", " ").replace("<ol>","").replace("</ol>","")\
            .replace("<li>","").replace("&lt;p>", "").replace("<ul>","").replace("</ul>","").replace("&lt;/p>", "")\
            .replace("1、", "1.").replace("2、", "2.").replace("3、", "3.").replace("4、", "4.").replace("5、", "5.")\
            .replace("6、", "6.").replace("&mdash;","").replace("']",'').replace("['",'').replace("\\n",'')
        #以上主要是基础替换

        #'''以下名称勿改'''
        # 以' '为参照进行字符串分割，分割完变成列表
        cut_apart = replace_character1.split("' '")
        # 取出列表第一条数据，为内容主体
        cut_apart1 = (cut_apart[0])
        # print(cut_apart1)
        # 用空替换['
        replace_character7 = cut_apart1.replace("['", "")
        # print(replace_character7)
        # 提取期望结果
        expectedresult = (cut_apart[3])

        # 获得最终期望结果，用空替换']
        expectedresult_last = expectedresult.replace("']", "")

        # 先将<!--替换成，普通字符l
        expectedresult_last2 = expectedresult_last.replace("<!--", "complex").replace("-->", "complex")
        # 再将-->替换成，普通字符l
        # 分组标定，替换，<!-- -->
        pattern=re.compile('(complex)(.*)(complex)')
        # 如果想包括两个1，则用pattern.sub(r\1''\3,content)
        expectedresult_last3 = pattern.sub(r'', expectedresult_last2)
        expectedresult_last4=expectedresult_last3.replace("<","complexx").replace(">","complexx")
        pattern1=re.compile('(complexx)(.*)(complexx)')
        expectedresult_last5=pattern1.sub(r'', expectedresult_last4)
        # print(expectedresult_last3)
        # print(expectedresult_last)
#         if cut_apart[2]=='none':
        module =cut_apart[4]
        # print(module)
             # 将所有模块层级添加到列表module_list
        module_list.append(module)
#         else:
             # 获取最终模块层级
#             module = replace_character7 + '/' + cut_apart[1] + '/' + cut_apart[2]
             # 将所有模块层级添加到列表module_list
#             module_list.append(module)
        # 将所有用例名称添加到列表testcase_list
        testcase_list.append(cut_apart[0])
        # 将所有用例前置条件添加到列表preconditions_list
        preconditions_list.append(cut_apart[1])
        # print(cut_apart[2])

        # 先将<!--替换成，普通字符l
        actions_last = cut_apart[2].replace("<!--", "complex").replace("-->", "complex")
        # 再将-->替换成，普通字符l
        # 分组标定，替换，<!-- -->
        pattern=re.compile('(complex)(.*)(complex)')
        # 如果想包括两个1，则用pattern.sub(r\1''\3,content)
        actions_last2 = pattern.sub(r'', actions_last)
        # print(actions_last2)

        actions_last3 = actions_last2.replace("<", "complexx").replace(">", "complexx")
        # 再将-->替换成，普通字符l
        # 分组标定，替换，<!-- -->
        pattern2=re.compile('(complexx)(.*)(complexx)')
        # 如果想包括两个1，则用pattern.sub(r\1''\3,content)
        actions_last4 = pattern2.sub(r'', actions_last3)
        # print(actions_last4)
        # 将所有用例操作步骤添加到列表actions_list
        actions_list.append(actions_last4)
        # 将所有用例期望结果添加到列表expectedresults_list
        expectedresults_list.append(expectedresult_last5)
    except Exception as e:
            if 'list index out of range' in str(e):
                print('数据读取错误，请检查！')
            else:
                print(str(e))

#保存为相应格式的xlsx文件
def set_sheet(sheet):
    sheet['A1'] = '用例名称'  # 加表头，给A1单元格赋值
    sheet['B1'] = '所属模块'  # 加表头，给B1单元格赋值
    sheet['C1'] = '用例类型'  # 加表头，给C1单元格赋值
    sheet['D1'] = '维护人'
    sheet['E1'] = '用例等级'
    sheet['F1'] = '测试方式'
    sheet['G1'] = '前置条件'
    sheet['H1'] = '备注'
    sheet['I1'] = '步骤描述'
    sheet['J1'] = '预期结果'

class XmlToXls(object):
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.create_sheet(sheet_name, 0)
        self.big_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        self.big_letter_num = []
        for i in self.big_letter:
            a = str(i) + '1'
            self.big_letter_num.append(a)

    def test_to_xls(self):
        try:
            self.wb = openpyxl.load_workbook(excel_name)
            try:
                del self.wb[sheet_name]
                self.__init__()
                set_sheet(self.sheet)
            except Exception as e:
                if 'does not exist' in str(e):
                    self.__init__()
                    set_sheet(self.sheet)
        except Exception as e:
            if 'No such file or directory' in str(e):
                self.__init__()
                set_sheet(self.sheet)

    def insert_data(self):
        for n in range(len(testcase_list)):
            self.sheet.append(
                [testcase_list[n], module_list[n], type_list[0], maintainer_list[0], testcase_level_list[0],
                 test_way_list[0], preconditions_list[n], comment_list[0], actions_list[n], expectedresults_list[n]])

        for i in self.big_letter_num:
            self.sheet[i].alignment = Alignment(horizontal='left', vertical='center')
        for m in range(1, 100):
            self.sheet.row_dimensions[m].height = 60
        testcase_path = os.path.abspath(os.path.join('metersphere_testcase'))
        # print(testcase_path)
        if not os.path.exists(testcase_path):
            os.makedirs(testcase_path)
        for i in self.big_letter:
            self.sheet.column_dimensions[i].width = 20
        # self.wb.save(excel_name)
        self.wb.save(os.path.abspath(os.path.join(testcase_path,excel_name)))


if __name__ == '__main__':
    a = XmlToXls()
    a.test_to_xls()
    a.insert_data()
